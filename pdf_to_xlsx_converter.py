#!/usr/bin/env python3
"""
PDF to XLSX Converter

A command-line utility to convert PDF files to XLSX format by extracting tables.
Scans the current working directory for PDF files and converts them to XLSX format.
"""

import os
import subprocess
import sys
from dataclasses import dataclass
from typing import List, Optional
import pandas as pd
import tabula
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Optional import for pdfplumber as backup
try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False


@dataclass
class ConversionResult:
    """Data class to track individual file conversion status"""
    file_name: str
    status: str  # 'success', 'failed', 'skipped'
    error_message: Optional[str] = None


@dataclass
class ConversionSummary:
    """Data class to track overall batch conversion results"""
    total_files: int
    successful: int
    failed: int
    skipped: int
    results: List[ConversionResult]


@dataclass
class TableExtractionResult:
    """Data class to track table extraction results"""
    success: bool
    tables: List[pd.DataFrame]
    error_message: Optional[str] = None
    extraction_method: str = "tabula"


def check_java_installation() -> tuple[bool, str]:
    """
    Check if Java is installed and accessible
    
    Returns:
        Tuple of (is_installed, version_info)
        - is_installed: True if Java is available
        - version_info: Java version string or error message
    """
    try:
        # Try to run java -version command
        result = subprocess.run(
            ['java', '-version'], 
            capture_output=True, 
            text=True, 
            timeout=10
        )
        
        if result.returncode == 0:
            # Java version info is typically in stderr
            version_output = result.stderr.strip()
            if version_output:
                # Extract version from output like "java version "1.8.0_291""
                lines = version_output.split('\n')
                if lines:
                    return True, lines[0]
            return True, "Java is installed"
        else:
            return False, "Java command failed to execute"
            
    except subprocess.TimeoutExpired:
        return False, "Java command timed out"
    except FileNotFoundError:
        return False, "Java command not found - Java may not be installed"
    except Exception as e:
        return False, f"Error checking Java installation: {e}"


def check_xlsx_dependencies() -> bool:
    """
    Check if all required dependencies for PDF to XLSX conversion are available
    
    Returns:
        True if all dependencies are satisfied, False otherwise
    """
    missing_deps = []
    
    # Check Java installation (required for tabula-py)
    java_installed, java_info = check_java_installation()
    if not java_installed:
        print(f"❌ Java Runtime Error: {java_info}")
        print("   tabula-py requires Java to be installed")
        print("   Please install Java from: https://www.java.com/download/")
        missing_deps.append("java")
    else:
        print(f"✅ Java check passed: {java_info}")
    
    # Check tabula-py
    try:
        import tabula
        print("✅ tabula-py is available")
    except ImportError:
        print("❌ tabula-py is not installed")
        print("   Please install it using: pip install tabula-py")
        missing_deps.append("tabula-py")
    
    # Check pandas
    try:
        import pandas as pd
        print("✅ pandas is available")
    except ImportError:
        print("❌ pandas is not installed")
        print("   Please install it using: pip install pandas")
        missing_deps.append("pandas")
    
    # Check openpyxl
    try:
        import openpyxl
        print("✅ openpyxl is available")
    except ImportError:
        print("❌ openpyxl is not installed")
        print("   Please install it using: pip install openpyxl")
        missing_deps.append("openpyxl")
    
    # Check pdfplumber (optional, used as fallback)
    try:
        import pdfplumber
        print("✅ pdfplumber is available (optional fallback)")
    except ImportError:
        print("⚠️ pdfplumber is not installed (optional)")
        print("   For better table extraction, install it using: pip install pdfplumber")
    
    if missing_deps:
        print(f"\n❌ Missing dependencies: {', '.join(missing_deps)}")
        print("\nTo install all required dependencies, run:")
        print("pip install tabula-py pandas openpyxl")
        if "java" in missing_deps:
            print("\nAlso install Java from: https://www.java.com/download/")
        return False
    
    print("\n✅ All dependencies are satisfied!")
    return True


def generate_xlsx_path(pdf_path: str) -> str:
    """
    Generate output XLSX file path from input PDF file path
    
    Args:
        pdf_path: Path to the input PDF file
        
    Returns:
        Path for the output XLSX file (same directory, same name, .xlsx extension)
    """
    # Get the directory and filename without extension
    directory = os.path.dirname(pdf_path)
    filename_without_ext = os.path.splitext(os.path.basename(pdf_path))[0]
    
    # Create the XLSX path
    xlsx_filename = f"{filename_without_ext}.xlsx"
    xlsx_path = os.path.join(directory, xlsx_filename)
    
    return xlsx_path


# Import existing utility functions from the original converter
# We'll reuse these functions from pdf_converter.py
def get_pdf_files(directory_path: str) -> List[str]:
    """
    Scan directory and identify PDF files
    (This will be imported/copied from the existing pdf_converter.py)
    """
    if not os.path.exists(directory_path):
        raise FileNotFoundError(f"Directory not found: {directory_path}")
    
    if not os.path.isdir(directory_path):
        raise NotADirectoryError(f"Path is not a directory: {directory_path}")
    
    try:
        pdf_files = []
        for file_name in os.listdir(directory_path):
            file_path = os.path.join(directory_path, file_name)
            
            # Check if it's a file (not a directory) and has .pdf extension
            if os.path.isfile(file_path) and file_name.lower().endswith('.pdf'):
                pdf_files.append(file_path)
        
        return sorted(pdf_files)  # Return sorted list for consistent ordering
        
    except PermissionError as e:
        raise PermissionError(f"Permission denied accessing directory: {directory_path}") from e


def is_valid_pdf_file(pdf_path: str) -> tuple[bool, str]:
    """
    Check if a file is a valid PDF by examining its structure
    (Copied from existing pdf_converter.py)
    """
    try:
        with open(pdf_path, 'rb') as f:
            # Read first 1024 bytes to check PDF structure
            header = f.read(1024)
            
            # Check for PDF header
            if not header.startswith(b'%PDF'):
                return False, "File does not have a valid PDF header"
            
            # Check file size first - empty or very small files are likely invalid
            f.seek(0, 2)  # Seek to end
            file_size = f.tell()
            if file_size < 100:  # PDF files should be at least 100 bytes
                return False, "File is too small to be a valid PDF"
            
            # Check for basic PDF structure markers
            if b'obj' not in header and b'endobj' not in header:
                # Read more of the file to look for object markers
                f.seek(0)
                content = f.read(4096)
                if b'obj' not in content:
                    return False, "File does not contain valid PDF object structure"
            
            # Look for EOF marker near the end of file
            f.seek(max(0, file_size - 1024))
            tail = f.read()
            if b'%%EOF' not in tail:
                return False, "File does not have a valid PDF end marker"
            
        return True, ""
        
    except Exception as e:
        return False, f"Error reading file: {e}"


def check_file_exists(file_path: str) -> bool:
    """
    Check if a file exists at the given path
    """
    return os.path.exists(file_path) and os.path.isfile(file_path)


def extract_tables_from_pdf(pdf_path: str) -> TableExtractionResult:
    """
    Extract all tables from a PDF file using tabula-py with multiple strategies
    
    Args:
        pdf_path: Path to the input PDF file
        
    Returns:
        TableExtractionResult containing extracted tables or error information
    """
    try:
        # Validate PDF file first
        is_valid, error_msg = is_valid_pdf_file(pdf_path)
        if not is_valid:
            return TableExtractionResult(
                success=False,
                tables=[],
                error_message=f"Invalid PDF file: {error_msg}",
                extraction_method="validation"
            )
        
        print(f"🔍 Extracting tables from: {os.path.basename(pdf_path)}")
        
        # Strategy 1: Try with default settings
        try:
            print("  📋 Trying default table extraction...")
            tables = tabula.read_pdf(
                pdf_path,
                pages='all',
                multiple_tables=True,
                pandas_options={'header': 0}
            )
            
            if tables and len(tables) > 0:
                valid_tables = process_extracted_tables(tables)
                if valid_tables:
                    return TableExtractionResult(
                        success=True,
                        tables=valid_tables,
                        error_message=None,
                        extraction_method="tabula-default"
                    )
        except Exception as e:
            print(f"  ⚠️ Default extraction failed: {e}")
        
        # Strategy 2: Try with lattice method (for tables with clear borders)
        try:
            print("  📋 Trying lattice method...")
            tables = tabula.read_pdf(
                pdf_path,
                pages='all',
                multiple_tables=True,
                lattice=True,
                pandas_options={'header': 0}
            )
            
            if tables and len(tables) > 0:
                valid_tables = process_extracted_tables(tables)
                if valid_tables:
                    return TableExtractionResult(
                        success=True,
                        tables=valid_tables,
                        error_message=None,
                        extraction_method="tabula-lattice"
                    )
        except Exception as e:
            print(f"  ⚠️ Lattice extraction failed: {e}")
        
        # Strategy 3: Try with stream method (for tables without clear borders)
        try:
            print("  📋 Trying stream method...")
            tables = tabula.read_pdf(
                pdf_path,
                pages='all',
                multiple_tables=True,
                stream=True,
                pandas_options={'header': 0}
            )
            
            if tables and len(tables) > 0:
                valid_tables = process_extracted_tables(tables)
                if valid_tables:
                    return TableExtractionResult(
                        success=True,
                        tables=valid_tables,
                        error_message=None,
                        extraction_method="tabula-stream"
                    )
        except Exception as e:
            print(f"  ⚠️ Stream extraction failed: {e}")
        
        # Strategy 4: Try to extract as raw text and structure it
        try:
            print("  📋 Trying text-based extraction...")
            # This is a fallback - extract all text and try to structure it
            tables = tabula.read_pdf(
                pdf_path,
                pages='all',
                multiple_tables=True,
                guess=False,  # Don't guess table areas
                pandas_options={'header': None}  # No header assumption
            )
            
            if tables and len(tables) > 0:
                valid_tables = process_extracted_tables(tables, assume_header=False)
                if valid_tables:
                    return TableExtractionResult(
                        success=True,
                        tables=valid_tables,
                        error_message=None,
                        extraction_method="tabula-text"
                    )
        except Exception as e:
            print(f"  ⚠️ Text extraction failed: {e}")
        
        # Strategy 5: Try pdfplumber as final fallback (if available)
        if PDFPLUMBER_AVAILABLE:
            try:
                print("  📋 Trying pdfplumber as fallback...")
                tables = extract_with_pdfplumber(pdf_path)
                
                if tables and len(tables) > 0:
                    valid_tables = process_extracted_tables(tables, assume_header=False)
                    if valid_tables:
                        return TableExtractionResult(
                            success=True,
                            tables=valid_tables,
                            error_message=None,
                            extraction_method="pdfplumber"
                        )
            except Exception as e:
                print(f"  ⚠️ pdfplumber extraction failed: {e}")
        
        # If all strategies failed
        return TableExtractionResult(
            success=False,
            tables=[],
            error_message="No tables found using any extraction method. The PDF may not contain tabular data or the tables may be in image format.",
            extraction_method="all-methods-failed"
        )
        
    except Exception as e:
        error_msg = str(e).lower()
        
        # Provide specific error messages for common issues
        if "java" in error_msg:
            detailed_error = "Java runtime error - ensure Java is properly installed"
        elif "memory" in error_msg:
            detailed_error = "Insufficient memory to process PDF"
        elif "timeout" in error_msg:
            detailed_error = "PDF processing timed out - file may be too complex"
        elif "permission" in error_msg:
            detailed_error = "Permission denied accessing PDF file"
        else:
            detailed_error = f"Table extraction failed: {e}"
        
        return TableExtractionResult(
            success=False,
            tables=[],
            error_message=detailed_error,
            extraction_method="tabula-error"
        )


def process_extracted_tables(tables: List[pd.DataFrame], assume_header: bool = True) -> List[pd.DataFrame]:
    """
    Process and clean a list of extracted tables
    
    Args:
        tables: List of raw DataFrames from table extraction
        assume_header: Whether to assume first row is header
        
    Returns:
        List of cleaned and valid DataFrames
    """
    valid_tables = []
    
    for i, table in enumerate(tables):
        if table is not None and not table.empty:
            # Clean the table data
            cleaned_table = clean_table_data(table, assume_header)
            if not cleaned_table.empty:
                valid_tables.append(cleaned_table)
                print(f"  ✅ Found table {i+1}: {cleaned_table.shape[0]} rows × {cleaned_table.shape[1]} columns")
    
    if valid_tables:
        print(f"  📊 Successfully extracted {len(valid_tables)} table(s)")
    
    return valid_tables


def extract_with_pdfplumber(pdf_path: str) -> List[pd.DataFrame]:
    """
    Extract tables using pdfplumber as a fallback method
    
    Args:
        pdf_path: Path to the PDF file
        
    Returns:
        List of DataFrames extracted from the PDF
    """
    if not PDFPLUMBER_AVAILABLE:
        return []
    
    tables = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                # Try to extract tables from the page
                page_tables = page.extract_tables()
                
                if page_tables:
                    for table_data in page_tables:
                        if table_data and len(table_data) > 1:  # At least 2 rows
                            # Convert to DataFrame
                            df = pd.DataFrame(table_data[1:], columns=table_data[0])
                            
                            # Clean the DataFrame
                            df = df.dropna(how='all')  # Remove empty rows
                            df = df.dropna(axis=1, how='all')  # Remove empty columns
                            
                            if not df.empty:
                                tables.append(df)
                
                # If no tables found, try to extract text and structure it
                if not page_tables:
                    text = page.extract_text()
                    if text:
                        # Try to find table-like structures in text
                        lines = text.split('\n')
                        table_lines = []
                        
                        for line in lines:
                            # Look for lines that might be table rows (contain multiple spaces or tabs)
                            if '\t' in line or '  ' in line:
                                # Split by tabs or multiple spaces
                                if '\t' in line:
                                    cells = line.split('\t')
                                else:
                                    cells = [cell.strip() for cell in line.split('  ') if cell.strip()]
                                
                                if len(cells) > 1:  # At least 2 columns
                                    table_lines.append(cells)
                        
                        # If we found table-like lines, create a DataFrame
                        if len(table_lines) > 1:
                            # Find the maximum number of columns
                            max_cols = max(len(row) for row in table_lines)
                            
                            # Pad rows to have the same number of columns
                            padded_lines = []
                            for row in table_lines:
                                padded_row = row + [''] * (max_cols - len(row))
                                padded_lines.append(padded_row)
                            
                            # Create DataFrame
                            if padded_lines:
                                df = pd.DataFrame(padded_lines[1:], columns=padded_lines[0])
                                df = df.dropna(how='all')
                                
                                if not df.empty:
                                    tables.append(df)
    
    except Exception as e:
        print(f"  ⚠️ pdfplumber processing error: {e}")
    
    return tables


def save_tables_to_xlsx(tables: List[pd.DataFrame], xlsx_path: str, source_pdf: str = "") -> None:
    """
    Save extracted tables to an XLSX file with proper formatting
    
    Args:
        tables: List of DataFrames containing table data
        xlsx_path: Path where the XLSX file should be saved
        source_pdf: Name of the source PDF file (for metadata)
        
    Raises:
        PermissionError: If file cannot be written due to permissions
        OSError: If there are disk space or other system issues
        Exception: For other Excel generation errors
    """
    try:
        # Create a new workbook
        wb = Workbook()
        
        # Remove the default sheet
        wb.remove(wb.active)
        
        # Define styles for formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Process each table
        for i, table in enumerate(tables):
            if table.empty:
                continue
                
            # Create worksheet name
            if len(tables) == 1:
                sheet_name = "Table"
            else:
                sheet_name = f"Table_{i+1}"
            
            # Ensure sheet name is valid (Excel has limitations)
            sheet_name = clean_sheet_name(sheet_name)
            
            # Create worksheet
            ws = wb.create_sheet(title=sheet_name)
            
            # Add table data to worksheet
            for r_idx, row in enumerate(dataframe_to_rows(table, index=False, header=True)):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx + 1, column=c_idx, value=value)
                    
                    # Apply border to all cells
                    cell.border = border
                    
                    # Apply header formatting to first row
                    if r_idx == 0:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                    
                    # Convert numeric strings to numbers for better Excel compatibility
                    if r_idx > 0 and isinstance(value, str) and value.strip():
                        try:
                            # Try to convert to float if it looks like a number
                            if value.replace('.', '').replace('-', '').replace('%', '').replace(',', '').isdigit():
                                if '%' in value:
                                    cell.value = float(value.replace('%', '')) / 100
                                    cell.number_format = '0.0%'
                                elif ',' in value:
                                    cell.value = float(value.replace(',', ''))
                                else:
                                    cell.value = float(value)
                        except (ValueError, AttributeError):
                            pass  # Keep as string if conversion fails
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width with reasonable limits
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add metadata as a comment or separate info
            if source_pdf:
                try:
                    from openpyxl.comments import Comment
                    ws['A1'].comment = Comment(f"Extracted from: {source_pdf}", "PDF Converter")
                except:
                    # If comment creation fails, just skip it
                    pass
        
        # If no valid tables were processed, create an info sheet
        if len(wb.worksheets) == 0:
            ws = wb.create_sheet(title="Info")
            ws['A1'] = "No valid tables found in the PDF file"
            ws['A2'] = f"Source: {source_pdf}" if source_pdf else "Source: Unknown"
        
        # Save the workbook
        wb.save(xlsx_path)
        print(f"  💾 Saved XLSX file: {os.path.basename(xlsx_path)}")
        
    except PermissionError as e:
        raise PermissionError(f"Permission denied writing XLSX file: {xlsx_path} - {e}")
    except OSError as e:
        if "space" in str(e).lower() or "disk" in str(e).lower():
            raise OSError(f"Insufficient disk space to save XLSX file: {xlsx_path}")
        else:
            raise OSError(f"System error saving XLSX file: {xlsx_path} - {e}")
    except Exception as e:
        raise Exception(f"Failed to create XLSX file: {e}")


def clean_sheet_name(name: str) -> str:
    """
    Clean sheet name to comply with Excel naming requirements
    
    Args:
        name: Proposed sheet name
        
    Returns:
        Valid Excel sheet name
    """
    # Excel sheet name limitations:
    # - Max 31 characters
    # - Cannot contain: \ / ? * [ ] :
    # - Cannot be empty
    
    if not name:
        name = "Sheet"
    
    # Remove invalid characters
    invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
    for char in invalid_chars:
        name = name.replace(char, '_')
    
    # Truncate to 31 characters
    if len(name) > 31:
        name = name[:31]
    
    return name


def validate_xlsx_output(xlsx_path: str) -> bool:
    """
    Validate that the generated XLSX file is valid and readable
    
    Args:
        xlsx_path: Path to the XLSX file to validate
        
    Returns:
        True if file is valid, False otherwise
    """
    try:
        # Check if file exists and has reasonable size
        if not os.path.exists(xlsx_path):
            return False
        
        file_size = os.path.getsize(xlsx_path)
        if file_size < 1000:  # XLSX files should be at least 1KB
            return False
        
        # Try to open the file with pandas to verify it's readable
        pd.read_excel(xlsx_path, sheet_name=0, nrows=1)
        return True
        
    except Exception:
        return False


def convert_pdf_to_xlsx(pdf_path: str, xlsx_path: str) -> bool:
    """
    Convert a PDF file to XLSX format by extracting tables
    
    Args:
        pdf_path: Path to the input PDF file
        xlsx_path: Path where the output XLSX file should be saved
        
    Returns:
        True if conversion was successful, False otherwise
        
    Raises:
        FileNotFoundError: If the PDF file doesn't exist
        PermissionError: If there are permission issues with file access
        ValueError: If the file is not a valid PDF
        Exception: For conversion errors
    """
    try:
        # Validate PDF file accessibility
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"PDF file not found: {pdf_path}")
        
        if not os.path.isfile(pdf_path):
            raise FileNotFoundError(f"Path is not a file: {pdf_path}")
        
        # Check if PDF file is readable and valid
        is_valid, error_msg = is_valid_pdf_file(pdf_path)
        if not is_valid:
            raise ValueError(f"Invalid PDF file: {error_msg}")
        
        # Check output directory permissions
        output_dir = os.path.dirname(xlsx_path)
        if output_dir and not os.path.exists(output_dir):
            raise FileNotFoundError(f"Output directory does not exist: {output_dir}")
        
        # Test write permissions
        try:
            temp_file = xlsx_path + ".tmp"
            with open(temp_file, 'w') as f:
                f.write("test")
            os.remove(temp_file)
        except PermissionError:
            raise PermissionError(f"Permission denied writing to output location: {xlsx_path}")
        except OSError as e:
            raise OSError(f"Cannot write to output location: {xlsx_path} - {e}")
        
        print(f"🔄 Converting: {os.path.basename(pdf_path)} → {os.path.basename(xlsx_path)}")
        
        # Extract tables from PDF
        extraction_result = extract_tables_from_pdf(pdf_path)
        
        if not extraction_result.success:
            raise Exception(f"Table extraction failed: {extraction_result.error_message}")
        
        if not extraction_result.tables:
            raise Exception("No tables found in the PDF file")
        
        # Save tables to XLSX
        save_tables_to_xlsx(
            extraction_result.tables, 
            xlsx_path, 
            os.path.basename(pdf_path)
        )
        
        # Validate the output file
        if not validate_xlsx_output(xlsx_path):
            raise Exception("Generated XLSX file failed validation")
        
        print(f"✅ Conversion successful: {os.path.basename(xlsx_path)}")
        print(f"   📊 Tables converted: {len(extraction_result.tables)}")
        print(f"   🔧 Extraction method: {extraction_result.extraction_method}")
        
        return True
        
    except Exception as e:
        # Clean up partial output file if it exists
        if os.path.exists(xlsx_path):
            try:
                os.remove(xlsx_path)
            except OSError:
                pass  # Ignore cleanup errors
        
        # Re-raise the original exception
        raise e


def should_convert_file(pdf_path: str, xlsx_path: str) -> bool:
    """
    Check if a file should be converted, handling overwrite protection
    
    Args:
        pdf_path: Path to the input PDF file
        xlsx_path: Path where the output XLSX file would be saved
        
    Returns:
        True if the file should be converted, False if it should be skipped
    """
    # If the XLSX file doesn't exist, proceed with conversion
    if not check_file_exists(xlsx_path):
        return True
    
    # If the XLSX file exists, prompt user for overwrite confirmation
    xlsx_filename = os.path.basename(xlsx_path)
    return prompt_overwrite(xlsx_filename)


def prompt_overwrite(file_name: str) -> bool:
    """
    Prompt user for confirmation on file overwrite
    
    Args:
        file_name: Name of the XLSX file that already exists
        
    Returns:
        True if user confirms overwrite, False if user declines
    """
    while True:
        try:
            response = input(f"⚠️  File '{file_name}' already exists. Overwrite? (y/n): ").strip().lower()
            
            if response in ['y', 'yes']:
                return True
            elif response in ['n', 'no']:
                return False
            else:
                print("Please enter 'y' for yes or 'n' for no.")
                
        except (EOFError, KeyboardInterrupt):
            # Handle Ctrl+C or EOF gracefully
            print("\n❌ Operation cancelled by user")
            return False


def handle_xlsx_conversion_error(error: Exception, file_name: str) -> None:
    """
    Handle XLSX conversion specific errors with detailed guidance
    
    Args:
        error: The exception that occurred during conversion
        file_name: Name of the file that caused the error
    """
    error_type = type(error).__name__
    error_message = str(error).lower()
    
    print(f"❌ Conversion Error for {file_name} ({error_type}):")
    
    if isinstance(error, ImportError):
        print(f"   {error}")
        if "tabula" in error_message:
            print("   Please install tabula-py: pip install tabula-py")
        elif "pandas" in error_message:
            print("   Please install pandas: pip install pandas")
        elif "openpyxl" in error_message:
            print("   Please install openpyxl: pip install openpyxl")
        else:
            print("   Please install required dependencies: pip install tabula-py pandas openpyxl")
    
    elif isinstance(error, FileNotFoundError):
        print(f"   {error}")
        if "java" in error_message:
            print("   Java Runtime Environment is required for tabula-py")
            print("   Please install Java from: https://www.java.com/download/")
        else:
            print("   The PDF file may have been moved or deleted during processing")
    
    elif isinstance(error, PermissionError):
        print(f"   {error}")
        print("   Check file permissions and ensure you have read/write access")
        print("   Try running with elevated permissions or check file ownership")
        if "xlsx" in error_message:
            print("   The XLSX file may be open in Excel - please close it and try again")
    
    elif isinstance(error, OSError):
        print(f"   {error}")
        if "disk" in error_message or "space" in error_message:
            print("   Insufficient disk space to complete the conversion")
            print("   Free up disk space and try again")
        elif "memory" in error_message:
            print("   Insufficient memory to process this PDF file")
            print("   Try closing other applications or processing smaller files")
        else:
            print("   A system-level error occurred during file processing")
    
    elif isinstance(error, ValueError):
        print(f"   {error}")
        if "pdf" in error_message:
            print("   The PDF file appears to be corrupted or invalid")
            print("   Try opening the file in a PDF viewer to verify its integrity")
        else:
            print("   Invalid data encountered during conversion")
    
    elif "java" in error_message:
        print(f"   {error}")
        print("   Java Runtime Error - tabula-py requires Java to be installed")
        print("   Please install Java from: https://www.java.com/download/")
        print("   Ensure Java is in your system PATH")
    
    elif "table" in error_message and "not found" in error_message:
        print(f"   {error}")
        print("   No tables were found in the PDF file")
        print("   This could mean:")
        print("     • The PDF contains only text without tabular data")
        print("     • The PDF is a scanned image (OCR may be needed)")
        print("     • Tables are in a format not recognized by the extraction tools")
        print("   Try using a different PDF or check if it contains actual tables")
    
    elif "memory" in error_message:
        print(f"   {error}")
        print("   The PDF file is too large or complex to process")
        print("   Try processing smaller files or increase available memory")
    
    elif "timeout" in error_message:
        print(f"   {error}")
        print("   PDF processing timed out - the file may be too complex")
        print("   Try processing a simpler PDF file")
    
    elif "excel" in error_message or "xlsx" in error_message:
        print(f"   {error}")
        print("   Error generating Excel file")
        print("   Check that the output directory is writable")
        print("   Ensure the XLSX file is not open in Excel")
    
    else:
        print(f"   {error}")
        print("   An unexpected error occurred during PDF to XLSX conversion")
        print("   Please check:")
        print("     • PDF file integrity")
        print("     • Available disk space")
        print("     • File permissions")
        print("     • Java installation (required for tabula-py)")
    
    print("   For more help, check the documentation or try with a different PDF file")
    print()


def display_progress(current: int, total: int, file_name: str) -> None:
    """
    Display current file being processed with progress information
    
    Args:
        current: Current file number being processed (1-based)
        total: Total number of files to process
        file_name: Name of the current file being processed
    """
    if total <= 0:
        print(f"🔄 Processing: {file_name}")
    else:
        percentage = (current / total) * 100
        print(f"🔄 Processing ({current}/{total} - {percentage:.0f}%): {file_name}")


def display_file_count(count: int) -> None:
    """
    Display the total number of PDF files found
    
    Args:
        count: Number of PDF files found in the directory
    """
    if count == 0:
        print("📁 No PDF files found in the current directory")
    elif count == 1:
        print("📁 Found 1 PDF file to convert")
    else:
        print(f"📁 Found {count} PDF files to convert")


def display_summary(successful: int, failed: int, skipped: int) -> None:
    """
    Display final conversion results summary
    
    Args:
        successful: Number of files successfully converted
        failed: Number of files that failed to convert
        skipped: Number of files that were skipped (due to existing XLSX files)
    """
    total = successful + failed + skipped
    
    print("\n" + "=" * 50)
    print("📊 CONVERSION SUMMARY")
    print("=" * 50)
    
    if total == 0:
        print("No files were processed")
        return
    
    print(f"Total files processed: {total}")
    
    if successful > 0:
        print(f"✅ Successfully converted: {successful}")
    
    if failed > 0:
        print(f"❌ Failed conversions: {failed}")
    
    if skipped > 0:
        print(f"⏭️  Skipped files: {skipped}")
    
    # Calculate and display success rate
    if total > 0:
        success_rate = (successful / total) * 100
        print(f"📈 Success rate: {success_rate:.1f}%")
    
    print("=" * 50)


def display_conversion_success(file_name: str) -> None:
    """
    Display success message for individual file conversion
    
    Args:
        file_name: Name of the file that was successfully converted
    """
    print(f"✅ Successfully converted: {file_name}")


def display_conversion_skipped(file_name: str) -> None:
    """
    Display message when a file is skipped due to existing XLSX
    
    Args:
        file_name: Name of the file that was skipped
    """
    print(f"⏭️  Skipped (file exists): {file_name}")


def pdf_to_xlsx_batch_convert() -> None:
    """
    Orchestrate the entire batch conversion process for PDF to XLSX
    
    This function integrates file discovery, conversion, error handling, and progress feedback.
    It continues processing after individual file failures and tracks conversion statistics.
    """
    # Initialize conversion statistics
    successful_count = 0
    failed_count = 0
    skipped_count = 0
    conversion_results = []
    
    try:
        # Get current working directory
        current_directory = os.getcwd()
        
        # Discover PDF files in the current directory
        pdf_files = get_pdf_files(current_directory)
        
        # Display total number of files found
        display_file_count(len(pdf_files))
        
        # If no PDF files found, exit early
        if not pdf_files:
            display_summary(successful_count, failed_count, skipped_count)
            return
        
        print()  # Add blank line for better readability
        
        # Process each PDF file
        for index, pdf_path in enumerate(pdf_files, 1):
            pdf_filename = os.path.basename(pdf_path)
            xlsx_path = generate_xlsx_path(pdf_path)
            xlsx_filename = os.path.basename(xlsx_path)
            
            # Display progress
            display_progress(index, len(pdf_files), pdf_filename)
            
            try:
                # Check if we should convert this file (handles overwrite protection)
                if not should_convert_file(pdf_path, xlsx_path):
                    # User declined to overwrite existing file
                    display_conversion_skipped(pdf_filename)
                    skipped_count += 1
                    conversion_results.append(ConversionResult(
                        file_name=pdf_filename,
                        status='skipped',
                        error_message='User declined to overwrite existing XLSX file'
                    ))
                    continue
                
                # Attempt the conversion
                success = convert_pdf_to_xlsx(pdf_path, xlsx_path)
                
                if success:
                    display_conversion_success(pdf_filename)
                    successful_count += 1
                    conversion_results.append(ConversionResult(
                        file_name=pdf_filename,
                        status='success'
                    ))
                else:
                    # This shouldn't happen with current implementation, but handle it
                    print(f"❌ Conversion failed for {pdf_filename}: Unknown error")
                    failed_count += 1
                    conversion_results.append(ConversionResult(
                        file_name=pdf_filename,
                        status='failed',
                        error_message='Unknown conversion error'
                    ))
                
            except Exception as error:
                # Handle conversion error gracefully and continue with next file
                handle_xlsx_conversion_error(error, pdf_filename)
                failed_count += 1
                conversion_results.append(ConversionResult(
                    file_name=pdf_filename,
                    status='failed',
                    error_message=str(error)
                ))
                # Continue processing the next file
                continue
        
        # Display final summary
        display_summary(successful_count, failed_count, skipped_count)
        
    except Exception as error:
        # Handle errors in file discovery or other critical failures
        print(f"\n❌ Critical error during batch conversion: {error}")
        print("Batch conversion process terminated.")
        
        # Still display summary of any files that were processed
        if successful_count > 0 or failed_count > 0 or skipped_count > 0:
            display_summary(successful_count, failed_count, skipped_count)


def clean_table_data(table: pd.DataFrame, assume_header: bool = True) -> pd.DataFrame:
    """
    Clean and validate extracted table data
    
    Args:
        table: Raw DataFrame from table extraction
        assume_header: Whether to assume first row is header
        
    Returns:
        Cleaned DataFrame
    """
    if table is None or table.empty:
        return pd.DataFrame()
    
    # Make a copy to avoid modifying the original
    cleaned = table.copy()
    
    # Remove completely empty rows and columns
    cleaned = cleaned.dropna(how='all')  # Remove rows where all values are NaN
    cleaned = cleaned.dropna(axis=1, how='all')  # Remove columns where all values are NaN
    
    # If table is still empty after cleaning, return empty DataFrame
    if cleaned.empty:
        return pd.DataFrame()
    
    # Reset index after dropping rows
    cleaned = cleaned.reset_index(drop=True)
    
    # Replace NaN values with empty strings for better Excel compatibility
    cleaned = cleaned.fillna('')
    
    # Clean column names - remove extra whitespace and handle unnamed columns
    new_columns = []
    for i, col in enumerate(cleaned.columns):
        if pd.isna(col) or str(col).strip() == '' or 'Unnamed' in str(col):
            new_columns.append(f'Column_{i+1}')
        else:
            new_columns.append(str(col).strip())
    cleaned.columns = new_columns
    
    # If we have very few rows (less than 2) and assume_header is True,
    # it might not be a real table
    if assume_header and len(cleaned) < 2:
        return pd.DataFrame()
    
    # Check if the table has meaningful content (not just empty strings)
    non_empty_cells = 0
    total_cells = cleaned.shape[0] * cleaned.shape[1]
    
    for col in cleaned.columns:
        for val in cleaned[col]:
            if str(val).strip() != '':
                non_empty_cells += 1
    
    # If less than 10% of cells have content, probably not a real table
    if total_cells > 0 and (non_empty_cells / total_cells) < 0.1:
        return pd.DataFrame()
    
    return cleaned


def main() -> None:
    """
    Main entry point for the PDF to XLSX converter application
    """
    try:
        print("PDF to XLSX Converter")
        print("=" * 50)
        print("Converting PDF files in current directory to XLSX format...")
        print()
        
        # Check dependencies before proceeding
        if not check_xlsx_dependencies():
            print("\n❌ Cannot proceed due to missing dependencies")
            print("Please install the required dependencies and try again.")
            return
        
        print("\n🎉 Dependencies check completed!")
        print("Ready to convert PDF files to XLSX format.")
        
        print()  # Add blank line after dependency check
        
        # Execute the main batch conversion process
        pdf_to_xlsx_batch_convert()
        
        print("\n🎉 Conversion process completed!")
        
    except KeyboardInterrupt:
        print("\n\n❌ Operation cancelled by user (Ctrl+C)")
        print("Conversion process terminated.")
        
    except Exception as error:
        print(f"\n❌ Unexpected application error: {error}")
        print("Please check your environment and try again.")
        
    finally:
        print("\nExiting PDF to XLSX Converter...")
        print("=" * 50)


if __name__ == "__main__":
    main()