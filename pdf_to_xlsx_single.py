#!/usr/bin/env python3
"""
PDF to XLSX Single File Converter

A command-line utility to convert a specific PDF file to XLSX format by extracting tables.
Usage: python pdf_to_xlsx_single.py <pdf_file_path> [output_file_path]
"""

import os
import sys
import subprocess
from typing import List, Optional
import pandas as pd
import tabula
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Optional import for pdfplumber as backup
try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False


def check_java_installation() -> tuple[bool, str]:
    """检查Java是否安装"""
    try:
        result = subprocess.run(
            ['java', '-version'], 
            capture_output=True, 
            text=True, 
            timeout=10
        )
        
        if result.returncode == 0:
            version_output = result.stderr.strip()
            if version_output:
                lines = version_output.split('\n')
                if lines:
                    return True, lines[0]
            return True, "Java is installed"
        else:
            return False, "Java command failed to execute"
            
    except subprocess.TimeoutExpired:
        return False, "Java command timed out"
    except FileNotFoundError:
        return False, "Java command not found - Java may not be installed"
    except Exception as e:
        return False, f"Error checking Java installation: {e}"


def check_dependencies():
    """检查所有必需的依赖"""
    missing_deps = []
    
    # 检查Java
    java_installed, java_info = check_java_installation()
    if not java_installed:
        print(f"❌ Java Runtime Error: {java_info}")
        print("   tabula-py需要Java运行环境")
        print("   请从以下地址安装Java: https://www.java.com/download/")
        missing_deps.append("java")
    else:
        print(f"✅ Java检查通过: {java_info}")
    
    # 检查Python库
    try:
        import tabula
        print("✅ tabula-py 可用")
    except ImportError:
        print("❌ 未安装tabula-py")
        print("   请使用以下命令安装: pip install tabula-py")
        missing_deps.append("tabula-py")
    
    try:
        import pandas as pd
        print("✅ pandas 可用")
    except ImportError:
        print("❌ 未安装pandas")
        print("   请使用以下命令安装: pip install pandas")
        missing_deps.append("pandas")
    
    try:
        import openpyxl
        print("✅ openpyxl 可用")
    except ImportError:
        print("❌ 未安装openpyxl")
        print("   请使用以下命令安装: pip install openpyxl")
        missing_deps.append("openpyxl")
    
    if PDFPLUMBER_AVAILABLE:
        print("✅ pdfplumber 可用 (可选备用)")
    else:
        print("⚠️ pdfplumber 未安装 (可选)")
        print("   为了更好的表格提取效果，建议安装: pip install pdfplumber")
    
    if missing_deps:
        print(f"\n❌ 缺少依赖: {', '.join(missing_deps)}")
        print("\n安装所有必需依赖:")
        print("pip install tabula-py pandas openpyxl")
        if "java" in missing_deps:
            print("\n同时需要安装Java: https://www.java.com/download/")
        return False
    
    print("\n✅ 所有依赖检查通过!")
    return True


def is_valid_pdf_file(pdf_path: str) -> tuple[bool, str]:
    """检查PDF文件是否有效"""
    try:
        with open(pdf_path, 'rb') as f:
            header = f.read(1024)
            
            if not header.startswith(b'%PDF'):
                return False, "文件没有有效的PDF头部"
            
            f.seek(0, 2)
            file_size = f.tell()
            if file_size < 100:
                return False, "文件太小，不是有效的PDF"
            
            if b'obj' not in header and b'endobj' not in header:
                f.seek(0)
                content = f.read(4096)
                if b'obj' not in content:
                    return False, "文件不包含有效的PDF对象结构"
            
            f.seek(max(0, file_size - 1024))
            tail = f.read()
            if b'%%EOF' not in tail:
                return False, "文件没有有效的PDF结束标记"
            
        return True, ""
        
    except Exception as e:
        return False, f"读取文件错误: {e}"


def extract_tables_from_pdf(pdf_path: str) -> tuple[bool, List[pd.DataFrame], str]:
    """从PDF中提取表格"""
    try:
        is_valid, error_msg = is_valid_pdf_file(pdf_path)
        if not is_valid:
            return False, [], f"无效的PDF文件: {error_msg}"
        
        print(f"🔍 从PDF中提取表格: {os.path.basename(pdf_path)}")
        
        # 策略1: 默认设置
        try:
            print("  📋 尝试默认表格提取...")
            tables = tabula.read_pdf(
                pdf_path,
                pages='all',
                multiple_tables=True,
                pandas_options={'header': 0}
            )
            
            if tables and len(tables) > 0:
                valid_tables = process_extracted_tables(tables)
                if valid_tables:
                    return True, valid_tables, "tabula-default"
        except Exception as e:
            print(f"  ⚠️ 默认提取失败: {e}")
        
        # 策略2: lattice方法
        try:
            print("  📋 尝试lattice方法...")
            tables = tabula.read_pdf(
                pdf_path,
                pages='all',
                multiple_tables=True,
                lattice=True,
                pandas_options={'header': 0}
            )
            
            if tables and len(tables) > 0:
                valid_tables = process_extracted_tables(tables)
                if valid_tables:
                    return True, valid_tables, "tabula-lattice"
        except Exception as e:
            print(f"  ⚠️ Lattice提取失败: {e}")
        
        # 策略3: stream方法
        try:
            print("  📋 尝试stream方法...")
            tables = tabula.read_pdf(
                pdf_path,
                pages='all',
                multiple_tables=True,
                stream=True,
                pandas_options={'header': 0}
            )
            
            if tables and len(tables) > 0:
                valid_tables = process_extracted_tables(tables)
                if valid_tables:
                    return True, valid_tables, "tabula-stream"
        except Exception as e:
            print(f"  ⚠️ Stream提取失败: {e}")
        
        # 策略4: pdfplumber备用方案
        if PDFPLUMBER_AVAILABLE:
            try:
                print("  📋 尝试pdfplumber备用方案...")
                tables = extract_with_pdfplumber(pdf_path)
                
                if tables and len(tables) > 0:
                    valid_tables = process_extracted_tables(tables, assume_header=False)
                    if valid_tables:
                        return True, valid_tables, "pdfplumber"
            except Exception as e:
                print(f"  ⚠️ pdfplumber提取失败: {e}")
        
        return False, [], "使用所有方法都未找到表格。PDF可能不包含表格数据或表格为图像格式。"
        
    except Exception as e:
        error_msg = str(e).lower()
        
        if "java" in error_msg:
            detailed_error = "Java运行时错误 - 确保Java已正确安装"
        elif "memory" in error_msg:
            detailed_error = "内存不足，无法处理PDF"
        elif "timeout" in error_msg:
            detailed_error = "PDF处理超时 - 文件可能过于复杂"
        elif "permission" in error_msg:
            detailed_error = "访问PDF文件权限被拒绝"
        else:
            detailed_error = f"表格提取失败: {e}"
        
        return False, [], detailed_error


def process_extracted_tables(tables: List[pd.DataFrame], assume_header: bool = True) -> List[pd.DataFrame]:
    """处理和清理提取的表格"""
    valid_tables = []
    
    for i, table in enumerate(tables):
        if table is not None and not table.empty:
            cleaned_table = clean_table_data(table, assume_header)
            if not cleaned_table.empty:
                valid_tables.append(cleaned_table)
                print(f"  ✅ 找到表格 {i+1}: {cleaned_table.shape[0]} 行 × {cleaned_table.shape[1]} 列")
    
    if valid_tables:
        print(f"  📊 成功提取 {len(valid_tables)} 个表格")
    
    return valid_tables


def extract_with_pdfplumber(pdf_path: str) -> List[pd.DataFrame]:
    """使用pdfplumber提取表格"""
    if not PDFPLUMBER_AVAILABLE:
        return []
    
    tables = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                page_tables = page.extract_tables()
                
                if page_tables:
                    for table_data in page_tables:
                        if table_data and len(table_data) > 1:
                            df = pd.DataFrame(table_data[1:], columns=table_data[0])
                            df = df.dropna(how='all')
                            df = df.dropna(axis=1, how='all')
                            
                            if not df.empty:
                                tables.append(df)
    
    except Exception as e:
        print(f"  ⚠️ pdfplumber处理错误: {e}")
    
    return tables


def clean_table_data(table: pd.DataFrame, assume_header: bool = True) -> pd.DataFrame:
    """清理表格数据"""
    if table is None or table.empty:
        return pd.DataFrame()
    
    cleaned = table.copy()
    
    # 删除完全空的行和列
    cleaned = cleaned.dropna(how='all')
    cleaned = cleaned.dropna(axis=1, how='all')
    
    if cleaned.empty:
        return pd.DataFrame()
    
    cleaned = cleaned.reset_index(drop=True)
    cleaned = cleaned.fillna('')
    
    # 清理列名
    new_columns = []
    for i, col in enumerate(cleaned.columns):
        if pd.isna(col) or str(col).strip() == '' or 'Unnamed' in str(col):
            new_columns.append(f'列_{i+1}')
        else:
            new_columns.append(str(col).strip())
    cleaned.columns = new_columns
    
    # 检查是否有意义的内容
    if assume_header and len(cleaned) < 2:
        return pd.DataFrame()
    
    non_empty_cells = 0
    total_cells = cleaned.shape[0] * cleaned.shape[1]
    
    for col in cleaned.columns:
        for val in cleaned[col]:
            if str(val).strip() != '':
                non_empty_cells += 1
    
    if total_cells > 0 and (non_empty_cells / total_cells) < 0.1:
        return pd.DataFrame()
    
    return cleaned


def save_tables_to_xlsx(tables: List[pd.DataFrame], xlsx_path: str, source_pdf: str = "") -> None:
    """保存表格到XLSX文件"""
    try:
        wb = Workbook()
        wb.remove(wb.active)
        
        # 定义样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # 处理每个表格
        for i, table in enumerate(tables):
            if table.empty:
                continue
                
            if len(tables) == 1:
                sheet_name = "表格"
            else:
                sheet_name = f"表格_{i+1}"
            
            sheet_name = clean_sheet_name(sheet_name)
            ws = wb.create_sheet(title=sheet_name)
            
            # 添加表格数据
            for r_idx, row in enumerate(dataframe_to_rows(table, index=False, header=True)):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx + 1, column=c_idx, value=value)
                    cell.border = border
                    
                    if r_idx == 0:  # 表头
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                    
                    # 数字格式转换
                    if r_idx > 0 and isinstance(value, str) and value.strip():
                        try:
                            if value.replace('.', '').replace('-', '').replace('%', '').replace(',', '').isdigit():
                                if '%' in value:
                                    cell.value = float(value.replace('%', '')) / 100
                                    cell.number_format = '0.0%'
                                elif ',' in value:
                                    cell.value = float(value.replace(',', ''))
                                else:
                                    cell.value = float(value)
                        except (ValueError, AttributeError):
                            pass
            
            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 如果没有有效表格，创建信息表
        if len(wb.worksheets) == 0:
            ws = wb.create_sheet(title="信息")
            ws['A1'] = "PDF文件中未找到有效表格"
            ws['A2'] = f"源文件: {source_pdf}" if source_pdf else "源文件: 未知"
        
        wb.save(xlsx_path)
        print(f"  💾 已保存XLSX文件: {os.path.basename(xlsx_path)}")
        
    except Exception as e:
        raise Exception(f"创建XLSX文件失败: {e}")


def clean_sheet_name(name: str) -> str:
    """清理工作表名称以符合Excel要求"""
    if not name:
        name = "工作表"
    
    invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
    for char in invalid_chars:
        name = name.replace(char, '_')
    
    if len(name) > 31:
        name = name[:31]
    
    return name


def convert_single_pdf_to_xlsx(pdf_path: str, output_path: str = None) -> bool:
    """转换单个PDF文件到XLSX格式"""
    # 检查PDF文件是否存在
    if not os.path.exists(pdf_path):
        print(f"❌ 错误: PDF文件不存在: {pdf_path}")
        return False
    
    # 如果没有指定输出路径，生成默认路径
    if output_path is None:
        base_name = os.path.splitext(pdf_path)[0]
        output_path = f"{base_name}.xlsx"
    
    # 检查输出文件是否已存在
    if os.path.exists(output_path):
        response = input(f"⚠️  文件 '{output_path}' 已存在，是否覆盖? (y/n): ").strip().lower()
        if response not in ['y', 'yes', '是']:
            print("❌ 转换已取消")
            return False
    
    print(f"🔄 开始转换: {os.path.basename(pdf_path)} → {os.path.basename(output_path)}")
    
    try:
        # 提取表格
        success, tables, method = extract_tables_from_pdf(pdf_path)
        
        if not success:
            print(f"❌ 表格提取失败: {method}")
            return False
        
        if not tables:
            print("❌ PDF中未找到表格")
            return False
        
        # 保存到XLSX
        save_tables_to_xlsx(tables, output_path, os.path.basename(pdf_path))
        
        print(f"✅ 转换成功!")
        print(f"   📊 转换的表格数: {len(tables)}")
        print(f"   🔧 提取方法: {method}")
        print(f"   📁 输出文件: {output_path}")
        return True
        
    except Exception as e:
        print(f"❌ 转换失败: {e}")
        # 清理可能的部分输出文件
        if os.path.exists(output_path):
            try:
                os.remove(output_path)
            except:
                pass
        return False


def main():
    """主函数"""
    print("PDF to XLSX 单文件转换器")
    print("=" * 40)
    
    # 检查依赖
    if not check_dependencies():
        return
    
    # 检查命令行参数
    if len(sys.argv) < 2:
        print("\n使用方法:")
        print(f"  python {os.path.basename(__file__)} <PDF文件路径> [输出文件路径]")
        print()
        print("示例:")
        print(f"  python {os.path.basename(__file__)} pdf2excel.pdf")
        print(f"  python {os.path.basename(__file__)} pdf2excel.pdf output.xlsx")
        return
    
    pdf_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    print(f"\n🎯 准备转换PDF文件: {pdf_file}")
    
    # 执行转换
    success = convert_single_pdf_to_xlsx(pdf_file, output_file)
    
    if success:
        print("\n🎉 转换完成!")
    else:
        print("\n❌ 转换失败!")
    
    print("=" * 40)


if __name__ == "__main__":
    main()